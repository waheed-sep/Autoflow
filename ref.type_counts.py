# import re
# from collections import Counter
# import pandas as pd
# from openpyxl import load_workbook
#
# # Function to extract and count occurrences of "type" values from a JSON file
# def extract_type_counts(file_path):
#     # Read the JSON file content as a string
#     with open(file_path, 'r') as file:
#         json_content = file.read()
#
#     # Regex pattern to find all "type" attribute values
#     pattern = r'"type"\s*:\s*"([^"]+)"'
#     matches = re.findall(pattern, json_content)
#
#     # Use Counter to count occurrences of each "type" value
#     type_counts = Counter(matches)
#
#     # Sort the dictionary by the values (occurrences) in descending order
#     sorted_type_counts = dict(sorted(type_counts.items(), key=lambda item: item[1], reverse=True))
#
#     return sorted_type_counts
#
# # Function to export results to an existing Excel file as a new worksheet
# def export_to_excel(data, output_file, sheet_name):
#     # Convert the dictionary to a DataFrame
#     df = pd.DataFrame(data.items(), columns=["Commit hash", "Occurrences"])
#
#     # Load the existing workbook
#     book = load_workbook(output_file)
#
#     with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
#         # Assign the existing workbook to the writer
#         writer._book = book
#
#         # Write the DataFrame to the specified sheet
#         df.to_excel(writer, sheet_name=sheet_name, index=False)
#
#         # Save the workbook
#         writer._save()
#
#     print(f"Results have been exported to '{sheet_name}' sheet in {output_file}")
#
# # Example usage
# file_path = './cloned_repo/PowerJoule/refactoring_results.json'  # Replace with the path to your JSON file
# type_counts = extract_type_counts(file_path)
#
# # Specify the output Excel file path and sheet name
# output_file = './cloned_repo/PowerJoule/Insight.xlsx'
# sheet_name = 'Commit_basics'
#
# # Export the results to an Excel file
# export_to_excel(type_counts, output_file, sheet_name)

import re
from collections import Counter
import pandas as pd
from openpyxl import load_workbook

# Function to extract and count occurrences of "type" values from a JSON file
def extract_type_counts(file_path):
    # Read the JSON file content as a string
    with open(file_path, 'r') as file:
        json_content = file.read()

    # Regex pattern to find all "type" attribute values
    pattern = r'"type"\s*:\s*"([^"]+)"'
    matches = re.findall(pattern, json_content)

    # Use Counter to count occurrences of each "type" value
    type_counts = Counter(matches)

    # Sort the dictionary by the values (occurrences) in descending order
    sorted_type_counts = dict(sorted(type_counts.items(), key=lambda item: item[1], reverse=True))

    return sorted_type_counts

# Function to export results to an existing Excel file as a new worksheet
def export_to_excel(data, output_file, sheet_name):
    # Convert the dictionary to a DataFrame
    df = pd.DataFrame(data.items(), columns=["Ref.type counts", "Occurrences"])

    # Load the existing workbook
    book = load_workbook(output_file)

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        # Associate the loaded workbook with the writer without setting writer.book
        writer._book = book  # This line may trigger the error, so let's remove direct assignment

        # Write the DataFrame to the specified sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Results have been exported to '{sheet_name}' sheet in {output_file}")

# Example usage
file_path = '../jflex/jflex_exp/jflex_refs_results.json'  # Replace with the path to your JSON file
type_counts = extract_type_counts(file_path)

# Specify the output Excel file path and sheet name
output_file = '../jflex/jflex_exp/jflex_insight.xlsx'
sheet_name = 'Ref.type counts'

# Export the results to an Excel file
export_to_excel(type_counts, output_file, sheet_name)
