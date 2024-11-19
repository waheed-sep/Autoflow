import subprocess
import pandas as pd
from openpyxl import Workbook

def run_script(script_name):
    """
    Run a Python script as a subprocess and capture its output.
    """
    subprocess.run(['python3', script_name], check=True)

def write_to_excel(sheet_name, data, workbook):
    """
    Write data to the specified sheet in the given workbook.
    """
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    sheet = workbook[sheet_name]
    for row in data:
        sheet.append(row)

def execute_and_store_results():
    """
    Executes all scripts in sequence and stores their results in a single Excel workbook.
    """
    # Initialize a workbook to store results
    workbook = Workbook()
    # Remove the default sheet created on workbook initialization
    workbook.remove(workbook.active)

    # Run each script and store its output into separate sheets in the workbook
    print("Running autoflow_main.py...")
    run_script('autoflow_main.py')
    # Assuming autoflow_main.py generates an output that can be read and written to Excel
    # Replace this with actual code to load the output from autoflow_main.py
    autoflow_data = [['Column1', 'Column2'], ['Data1', 'Data2']]  # Example data
    write_to_excel('Autoflow Main', autoflow_data, workbook)

    print("Running commits_insights.py...")
    run_script('commits_insights.py')
    # Replace this with actual code to load the output from commits_insights.py
    commits_data = [['CommitID', 'Message'], ['abc123', 'Fixed bug']]  # Example data
    write_to_excel('Commits Insights', commits_data, workbook)

    print("Running ref.type_counts.py...")
    run_script('ref.type_counts.py')
    # Replace this with actual code to load the output from ref.type_counts.py
    ref_data = [['RefType', 'Count'], ['Type1', 5]]  # Example data
    write_to_excel('Ref Type Counts', ref_data, workbook)

    # Save the workbook to a single file
    workbook.save('Insights.xlsx')
    print("All scripts executed successfully. Results saved to 'Insights.xlsx'.")

if __name__ == '__main__':
    execute_and_store_results()




# import subprocess
# import pandas as pd
#
# # Function to run a script and capture its output
# def run_script(script_name):
#     try:
#         result = subprocess.run(['python3', 'autoflow_main.py'], capture_output=True, text=True, check=True)
#         return result.stdout
#     except subprocess.CalledProcessError as e:
#         print(f"Error occurred while running {script_name}: {e}")
#         return None
#
# # Run the Python scripts and capture their output
# commits_output = run_script('commits_insights.py')
# refs_output = run_script('ref.type_counts.py')
#
# def parse_output_to_dataframe(output):
#     if output:
#         # Assuming output is tabular (e.g., CSV format)
#         data = [line.split() for line in output.strip().split('\\n')]
#         columns = data[0]  # First row as headers
#         rows = data[1:]    # Subsequent rows as data
#         return pd.DataFrame(rows, columns=columns)
#     else:
#         return pd.DataFrame()
#
# # Convert script outputs to DataFrames
# commits_df = parse_output_to_dataframe(commits_output)
# refs_df = parse_output_to_dataframe(refs_output)
#
# # Create an Excel writer
# with pd.ExcelWriter('Insights.xlsx', engine='openpyxl') as writer:
#     if not commits_df.empty:
#         commits_df.to_excel(writer, sheet_name='Commits insights', index=False)
#     if not refs_df.empty:
#         refs_df.to_excel(writer, sheet_name='Refs count', index=False)
#
# # You can add more DataFrames or outputs here if needed
